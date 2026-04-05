import pandas as pd
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from pandas.testing import assert_frame_equal

from tabflow import PageRange


def make_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report Sheet"
    return wb, ws


def add_named_range(wb, name: str, reference: str) -> None:
    wb.defined_names.add(DefinedName(name, attr_text=reference))


def test_update_single_cell_updates_cell_value():
    wb, ws = make_workbook()
    ws["A1"] = "Old value"
    add_named_range(wb, "PERIOD", "'Report Sheet'!$A$1")

    PageRange(wb, "PERIOD").update_single_cell("New value")

    assert ws["A1"].value == "New value"


def test_get_data_returns_dataframe():
    wb, ws = make_workbook()
    ws.append(["Store", "Sales"])
    ws.append(["Store 1", 42])
    add_named_range(wb, "DATA", "'Report Sheet'!$A$1:$B$2")

    result = PageRange(wb, "DATA").get_data()

    expected = pd.DataFrame([
        ["Store", "Sales"],
        ["Store 1", 42],
    ])
    assert_frame_equal(result, expected)


def test_set_data_with_header_expands_named_range():
    wb, ws = make_workbook()
    ws.append(["Store", "Sales"])
    ws.append(["Old Store", 1])
    add_named_range(wb, "DATA", "'Report Sheet'!$A$1:$B$2")

    replacement = pd.DataFrame(
        [["Store 1", 10], ["Store 2", 20]],
        columns=["Store", "Sales"],
    )

    page_range = PageRange(wb, "DATA")
    page_range.set_data(replacement, has_header=True)

    assert page_range.cell_range == "A1:B3"
    assert wb.defined_names["DATA"].attr_text == "'Report Sheet'!$A$1:$B$3"
    assert ws["A2"].value == "Store 1"
    assert ws["B3"].value == 20


def test_clear_data_keeps_header_when_requested():
    wb, ws = make_workbook()
    ws.append(["Store", "Sales"])
    ws.append(["Store 1", 42])
    ws.append(["Store 2", 21])
    add_named_range(wb, "DATA", "'Report Sheet'!$A$1:$B$3")

    PageRange(wb, "DATA").clear_data(keep_header=True)

    assert ws["A1"].value == "Store"
    assert ws["B1"].value == "Sales"
    assert ws["A2"].value is None
    assert ws["B3"].value is None
