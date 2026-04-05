from __future__ import annotations

from copy import copy
from typing import Any

import pandas as pd
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import get_column_letter


class PageRange:
    """Read and update an Excel named range while preserving formatting"""

    def __init__(self, workbook, name: str):
        """Initialize an object that manages a workbook named range"""
        self.workbook = workbook
        self.name = name
        self.sheet_name, self.cell_range = self._parse_named_range(name)
        self.data = pd.DataFrame()

    def _parse_named_range(self, name: str) -> tuple[str, str]:
        """Return ``(sheet_name, cell_range)`` for a workbook defined name"""
        defined_name = self.workbook.defined_names.get(name)
        if defined_name is None or not getattr(defined_name, "attr_text", None):
            raise ValueError(f"Named range '{name}' not found in workbook")

        try:
            sheet_ref, cell_ref = defined_name.attr_text.split("!", 1)
        except ValueError as exc:
            raise ValueError(
                f"Named range '{name}' has invalid reference: {defined_name.attr_text}"
            ) from exc

        sheet_name = sheet_ref.strip("'").replace("''", "'")
        cell_range = cell_ref.replace("$", "")
        return sheet_name, cell_range

    def _quoted_sheet_name(self) -> str:
        """Return sheet name quoted only when Excel references require it"""
        if " " in self.sheet_name or "-" in self.sheet_name:
            escaped_name = self.sheet_name.replace("'", "''")
            return f"'{escaped_name}'"
        return self.sheet_name

    def _normalize_data_input(self, new_data: Any) -> list[list[Any]]:
        """Convert supported input formats into a list of row lists"""
        if isinstance(new_data, pd.DataFrame):
            return new_data.values.tolist()

        if isinstance(new_data, list):
            if not new_data:
                return []

            first_row = new_data[0]
            if isinstance(first_row, dict):
                return [list(row.values()) for row in new_data]

            if isinstance(first_row, (list, tuple)):
                return [list(row) for row in new_data]

        raise ValueError("new_data must be a DataFrame, list of lists, or list of dicts")

    def _set_local_range(self, min_col: int, min_row: int, max_col: int, max_row: int) -> None:
        """Update the cached local range string"""
        self.cell_range = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )

    def get_data(self) -> pd.DataFrame:
        """Retrieve data from the named range as a pandas DataFrame"""
        min_col, min_row, max_col, max_row = range_boundaries(self.cell_range)
        ws = self.workbook[self.sheet_name]

        data = []
        for row in range(min_row, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                row_data.append(ws.cell(row=row, column=col).value)
            data.append(row_data)

        self.data = pd.DataFrame(data)
        return self.data

    def _copy_cell_format(self, source_cell, target_cell) -> None:
        """Copy all formatting from source cell to target cell"""
        if not source_cell.has_style:
            return

        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

    def _get_current_named_range_bounds(self):
        """Get current boundaries of the named range from workbook metadata"""
        self.sheet_name, self.cell_range = self._parse_named_range(self.name)
        return range_boundaries(self.cell_range)

    def _update_named_range(self, min_col: int, min_row: int, max_col: int, max_row: int) -> None:
        """Update named range definition and local cached range string"""
        col_start = get_column_letter(min_col)
        col_end = get_column_letter(max_col)
        sheet_ref = self._quoted_sheet_name()
        new_range = f"{sheet_ref}!${col_start}${min_row}:${col_end}${max_row}"

        self.workbook.defined_names[self.name].attr_text = new_range
        self._set_local_range(min_col, min_row, max_col, max_row)

    def _expand_conditional_formatting(
        self,
        ws,
        min_col: int,
        min_row: int,
        max_col: int,
        old_max_row: int,
        new_max_row: int,
    ) -> None:
        """Expand conditional formatting rules to cover newly inserted rows"""
        if not hasattr(ws, "conditional_formatting"):
            return

        cf_to_update = []

        for cf_obj, rule_list in list(ws.conditional_formatting._cf_rules.items()):
            old_sqref = str(cf_obj.sqref)

            for cf_range_str in old_sqref.split():
                try:
                    cf_min_col, cf_min_row, cf_max_col, cf_max_row = range_boundaries(cf_range_str)
                except ValueError:
                    continue

                col_overlaps = not (cf_max_col < min_col or cf_min_col > max_col)
                row_affected = cf_max_row >= min_row and cf_max_row <= old_max_row
                needs_expansion = cf_max_row < new_max_row

                if col_overlaps and row_affected and needs_expansion:
                    new_range = (
                        f"{get_column_letter(cf_min_col)}{cf_min_row}:"
                        f"{get_column_letter(cf_max_col)}{new_max_row}"
                    )
                    cf_to_update.append((cf_obj, rule_list, new_range))
                    break

        for cf_obj, rule_list, new_range in cf_to_update:
            del ws.conditional_formatting._cf_rules[cf_obj]
            for rule in rule_list:
                ws.conditional_formatting.add(new_range, rule)

    def set_data(self, new_data: Any, has_header: bool = False) -> None:
        """
        Update the named range with new data while preserving formatting.
        Handles row expansion/shrinking before writing values.
        """
        data_list = self._normalize_data_input(new_data)
        if not data_list:
            return

        min_col, min_row, max_col, max_row = range_boundaries(self.cell_range)
        ws = self.workbook[self.sheet_name]

        data_start_row = min_row + 1 if has_header else min_row
        original_data_rows = max_row - data_start_row + 1
        new_data_rows = len(data_list)
        rows_difference = new_data_rows - original_data_rows

        old_max_row = max_row
        template_row = data_start_row
        template_cells = [
            ws.cell(row=template_row, column=col)
            for col in range(min_col, max_col + 1)
        ]

        if rows_difference > 0:
            ws.insert_rows(max_row + 1, rows_difference)
            desired_new_max_row = data_start_row + new_data_rows - 1
            updated_bounds = self._get_current_named_range_bounds()

            if updated_bounds[3] != desired_new_max_row:
                self._update_named_range(min_col, min_row, max_col, desired_new_max_row)
            else:
                self._set_local_range(min_col, min_row, max_col, desired_new_max_row)

            new_max_row = desired_new_max_row

        elif rows_difference < 0:
            delete_position = data_start_row + new_data_rows
            ws.delete_rows(delete_position, abs(rows_difference))
            new_max_row = data_start_row + new_data_rows - 1
            updated_bounds = self._get_current_named_range_bounds()

            if updated_bounds[3] != new_max_row:
                self._update_named_range(min_col, min_row, max_col, new_max_row)
            else:
                self._set_local_range(min_col, min_row, max_col, new_max_row)
        else:
            new_max_row = max_row

        data_width = max_col - min_col + 1
        for row_idx, row_data in enumerate(data_list):
            current_row = data_start_row + row_idx

            for col_idx, value in enumerate(row_data[:data_width]):
                current_col = min_col + col_idx
                cell = ws.cell(row=current_row, column=current_col)
                cell.value = value
                self._copy_cell_format(template_cells[col_idx], cell)

        if rows_difference > 0:
            self._expand_conditional_formatting(
                ws, min_col, min_row, max_col, old_max_row, new_max_row
            )

    def update_single_cell(self, value: Any) -> None:
        """Update a single-cell named range while preserving formatting"""
        min_col, min_row, max_col, max_row = range_boundaries(self.cell_range)

        if min_row != max_row or min_col != max_col:
            raise ValueError(f"Named range '{self.name}' is not a single cell")

        ws = self.workbook[self.sheet_name]
        ws.cell(row=min_row, column=min_col).value = value

    def clear_data(self, keep_header: bool = False) -> None:
        """Clear values from the named range while preserving formatting"""
        min_col, min_row, max_col, max_row = range_boundaries(self.cell_range)
        ws = self.workbook[self.sheet_name]

        start_row = min_row + 1 if keep_header else min_row
        for row in range(start_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = None


__all__ = ["PageRange"]
